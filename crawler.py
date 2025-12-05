import os
import shutil
import time
import datetime
import logging
import requests
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment

def setup_logger():
    logging.basicConfig(
        level=logging.INFO,  # INFO 레벨로 설정
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(),  # 콘솔 출력
            logging.FileHandler("crawler.log")  # 로그 파일
        ]
    )

def click_page_button(driver, nth_child):
    try:
        page_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, f".pagination > a:nth-child({nth_child})"))
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", page_button)
        time.sleep(0.5)
        page_button.click()
        # 페이지 로드 대기
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".room_item"))
        )
        time.sleep(2)  # 추가 페이지 로드 대기
        logging.info(f"이 섹션의 {nth_child - 1}번째 페이지 버튼을 성공적으로 클릭했습니다.")
    except Exception as e:
        logging.error(f"페이지 버튼 {nth_child} 클릭 중 오류 발생: {e}")
        raise

def click_next_section(driver):
    try:
        next_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".pagination > .next.is_active"))
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
        time.sleep(0.5)
        next_button.click()
        # 섹션 로드 대기
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".room_item"))
        )
        time.sleep(3)  # 추가 로드 대기
        logging.info("다음 섹션으로 이동했습니다.")
    except Exception as e:
        logging.error(f"다음 섹션 버튼 클릭 중 오류 발생: {e}")
        raise

def test_pagination(driver, data_list, image_dir, max_sections=100, pages_per_section=10):
    section_number = 1  # 섹션 번호 초기화

    while section_number <= max_sections:
        
        try:
            if section_number == 1:
                # 첫 번째 섹션: 페이지 2는 a:nth-child(2), 페이지 3~12는 a:nth-child(4) ~ a:nth-child(11)
                click_page_button(driver, 2)  # 페이지 2
                process_rooms(driver, data_list, image_dir)  # 페이지 2 크롤링

                for nth in range(4, 12):  # 페이지 3~10
                    click_page_button(driver, nth)
                    process_rooms(driver, data_list, image_dir)  # 각 페이지 크롤링
            else:
                # 두 번째 섹션 이상: 페이지 12, 22, ...는 a:nth-child(3), 페이지 13~22, 23~32, ...는 a:nth-child(4) ~ a:nth-child(11)
                click_page_button(driver, 3)  # 두번째 페이지 (12, 22, 32, ...)
                process_rooms(driver, data_list, image_dir)  # 첫 페이지 크롤링

                for nth in range(4, 12):  # 페이지 다음 8개 (3번째~10번째 페이지)
                    click_page_button(driver, nth)
                    process_rooms(driver, data_list, image_dir)  # 각 페이지 크롤링

            # 다음 섹션으로 이동
            click_next_section(driver)
            # 다음 섹션으로 이동하자마자 바로 첫페이지가 나오기때문에 여기도 크롤링하고넘어가야함
            process_rooms(driver, data_list, image_dir)
            section_number += 1

        except Exception as e:
            logging.warning(f"섹션 {section_number} 처리 중 오류 발생 또는 마지막 페이지: {e}")
            break  # 오류 발생 시 페이지네이션 테스트 종료

    logging.info("페이지 넘기기 기능이 성공적으로 작동합니다.")

def get_total_pages(driver):

    try:
        pagination = driver.find_element(By.CSS_SELECTOR, ".pagination")
        page_buttons = pagination.find_elements(By.TAG_NAME, "a")
        page_numbers = []
        for btn in page_buttons:
            text = btn.text.strip()
            if text.isdigit():
                page_numbers.append(int(text))
        return max(page_numbers) if page_numbers else 1
    except Exception as e:
        logging.warning(f"총 페이지 수를 확인할 수 없습니다: {e}")
        return 1

def process_rooms(driver, data_list, image_dir):
    try:
        # 모든 방 링크 찾기
        room_links = driver.find_elements(By.CSS_SELECTOR, ".result_room > a")
        links = [link.get_attribute('href') for link in room_links]

        # 썸네일 추출
        room_elements = driver.find_elements(By.CSS_SELECTOR, ".room_item")
        thumbnails = []
        for idx, room in enumerate(room_elements, start=1):
            try:
                thumbnail_url = room.find_element(By.CSS_SELECTOR, ".room_item > dt > img").get_attribute('src')
                thumbnails.append(thumbnail_url)
            except Exception as e:
                logging.warning(f"썸네일 추출 오류: {e}")
                thumbnails.append(None)

        # 링크와 썸네일 수가 일치하는지 확인
        if len(links) != len(thumbnails):
            logging.warning("링크 수와 썸네일 수가 일치하지 않습니다.")

        # 각 게시물을 순회하며 데이터 수집
        for idx, (thumbnail_url, link) in enumerate(zip(thumbnails, links), start=1):
            try:
                logging.info(f"게시물 처리 중 {idx}/{len(links)}: {link}")

                # 새 탭에서 링크 열기
                driver.execute_script("window.open(arguments[0]);", link)
                driver.switch_to.window(driver.window_handles[1])

                # 페이지 로드 대기
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".room_detail"))
                )
                time.sleep(1)  # 추가 로드 대기

                # 방 상세 정보 수집
                title = driver.find_element(By.CSS_SELECTOR, "body > div.wrap > section > div > div.room_detail > div:nth-child(1) > div.title > strong").text
                address = driver.find_element(By.CSS_SELECTOR, "body > div.wrap > section > div > div.room_detail > div:nth-child(1) > p").text
                area = driver.find_element(By.CSS_SELECTOR, ".place_detail > li:nth-child(1) > strong").text
                type_of_room = driver.find_element(By.CSS_SELECTOR, ".place_detail > li:nth-child(2) > strong").text
                weekly_rent_price = driver.find_element(By.CSS_SELECTOR, ".tbl_style > tbody > tr > td:nth-child(1)").text
                management_price = driver.find_element(By.CSS_SELECTOR, ".tbl_style > tbody > tr > td:nth-child(2)").text
                cleaning_price = driver.find_element(By.CSS_SELECTOR, ".tbl_style > tbody > tr > td:nth-child(3)").text

                # 예약 확인 버튼 클릭
                reservation_check_btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "#btn_check_schdule"))
                )
                reservation_check_btn.click()
                time.sleep(1)  # 예약 상태 로드 대기

                # 예약 데이터 초기화
                reservation_data = {}
                total_disabled = 0
                total_total = 0
                current_month = datetime.datetime.now().month

                for j in range(1, 4):
                    try:
                        # 페이지 완전히 로드될 때까지 대기
                        WebDriverWait(driver, 10).until(
                            lambda d: d.execute_script('return document.readyState') == 'complete'
                        )

                        # 예약 상태 추출
                        html = driver.page_source
                        soup = BeautifulSoup(html, 'html.parser')

                        month_disabled = len(soup.select('.calendar_table > thead > tr > .disable'))
                        month_enabled = len(soup.select(".calendar_table > thead > tr > .enable"))
                        month_total = month_disabled + month_enabled

                        # 총합 누적
                        total_disabled += month_disabled
                        total_total += month_total

                        # 월 번호 계산 (12월 이후에는 1월로 돌아감)
                        month_num = (current_month + j - 1) % 12
                        month_num = 12 if month_num == 0 else month_num

                        # 예약 상태 저장
                        reservation_data[f"{month_num}월 예약"] = f"{month_disabled}/{month_total}"
                        logging.info(f"{month_num}월 예약현황 : {month_disabled} / {month_total}")

                        # 다음 달 버튼 클릭
                        next_month_btn = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, "#btn_next_month > img"))
                        )
                        driver.execute_script("arguments[0].scrollIntoView(true);", next_month_btn)
                        time.sleep(0.5)
                        actions = ActionChains(driver)
                        actions.move_to_element(next_month_btn).click().perform()

                        # 다음 달 페이지 로드 대기
                        WebDriverWait(driver, 10).until(
                            lambda d: d.execute_script('return document.readyState') == 'complete'
                        )
                        time.sleep(1)  # 추가 로드 대기

                    except Exception as e:
                        logging.warning(f"예약 데이터 수집 오류: {e}")
                        break  # 오류 발생 시 루프 종료

                # 예약률 계산
                reservation_rate = (total_disabled / total_total) if total_total > 0 else 0
                logging.info(f"평균 예약률: {reservation_rate * 100:.1f}%")

                # 데이터 딕셔너리 준비
                data = {
                    "대표이미지": thumbnail_url,
                    "매물명": title,
                    "주소": address,
                    "건물유형": type_of_room,
                    "전용면적": area,
                    "임대료(1주)": weekly_rent_price,
                    "관리비용": management_price,
                    "청소비용": cleaning_price,
                    "URL": link  # 'URL'을 키로 사용
                }

                # 예약 데이터와 예약률 추가
                data.update(reservation_data)
                data["예약률"] = f"{reservation_rate:.2%}"

                # 데이터 리스트에 추가
                data_list.append(data)

                logging.info(f"데이터 추가됨: {title}")

                # 현재 탭 닫고 원래 탭으로 전환
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                time.sleep(0.5)

            except Exception as e:
                logging.error(f"게시물 처리 중 오류: {e}", exc_info=True)
                # 오류 발생 시 현재 탭 닫고 원래 탭으로 전환
                if len(driver.window_handles) > 1:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                continue

    except Exception as e:
        logging.error(f"process_rooms 함수 오류: {e}", exc_info=True)

def crawl(keywords, base_image_dir, output_dir, max_sections=100, pages_per_section=10):
    setup_logger()

    # 크롬 드라이버 옵션
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--no-sandbox')

    # 드라이버 1번만 생성
    service = Service(ChromeDriverManager().install())
    service.log_path = os.devnull
    driver = webdriver.Chrome(service=service, options=chrome_options)

    try:
        # 로그인 1번만 수행
        driver.get("https://33m2.co.kr/")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".room_item"))
        )
        logging.info("웹사이트 접속 성공")
        time.sleep(2)
        logging.info("1분간 로그인 대기…")
        time.sleep(60)

        # ----------------------------
        # 🔥 키워드별로 따로 처리 시작
        # ----------------------------
        for keyword in keywords:

            logging.info(f"키워드 '{keyword}' 크롤링 시작")

            # 키워드별 이미지 디렉토리
            image_dir = os.path.join(base_image_dir, keyword)
            os.makedirs(image_dir, exist_ok=True)

            # 키워드별 엑셀 파일 경로
            excel_path = os.path.join(output_dir, f"rooms_data_{keyword}.xlsx")

            # 첫 번째 키워드는 로그인 직후 검색
            # 두 번째 이후는 guest 페이지로 돌아가 검색
            driver.get("https://33m2.co.kr")
            time.sleep(2)

            search = driver.find_element(By.CSS_SELECTOR, '#txt_search_keyword')
            search.clear()
            search.send_keys(keyword)

            driver.find_element(By.CSS_SELECTOR, "#btn_search").click()

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".room_item"))
            )
            time.sleep(2)

            # 결과 데이터 담을 리스트
            data_list = []

            # 페이지 1
            process_rooms(driver, data_list, image_dir)

            # 페이지네이션
            test_pagination(driver, data_list, image_dir, max_sections, pages_per_section)

            logging.info(f"키워드 '{keyword}' 크롤링 완료 → 엑셀 생성 시작")

            # ---------------------------
            # 🔥 엑셀 저장 (네 코드 그대로 유지)
            # ---------------------------
            if data_list:
                df = pd.DataFrame(data_list)
                df.insert(0, "순번", range(1, len(df) + 1))

                df.to_excel(excel_path, index=False, engine='openpyxl')

                wb = load_workbook(excel_path)
                ws = wb.active

                # 열 너비 조정
                column_widths = {
                    'A': 10,  # 순번
                    'B': 20,  # 대표이미지
                    'C': 30,  # 매물명
                    'D': 50,  # 주소
                    'E': 15,  # 건물유형
                    'F': 15,  # 전용면적
                    'G': 15,  # 임대료(1주)
                    'H': 15,  # 관리비용
                    'I': 15,  # 청소비용
                    'J': 50,  # URL
                    'K': 15,  # 첫번째달예약
                    'L': 15,  # 두번째달예약
                    'M': 15,  # 세번째달예약
                    'N': 15   # 예약률
                }

                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width

                # 이미지 삽입 및 셀 서식 지정
                for row in range(2, ws.max_row + 1):
                    # 대표이미지 삽입
                    img_url = ws.cell(row=row, column=2).value
                    if img_url:
                        try:
                            img_response = requests.get(img_url, stream=True)
                            img_path = os.path.join(image_dir, f"img_{row}.jpg")
                            with open(img_path, 'wb') as out_file:
                                shutil.copyfileobj(img_response.raw, out_file)
                            img = ExcelImage(img_path)
                            img.width = 155  # 필요에 따라 조정
                            img.height = 100  # 필요에 따라 조정
                            img.anchor = f'B{row}'
                            ws.add_image(img)

                            ws.cell(row=row, column=2).value = None  # 이미지 삽입 후 URL 제거
                        except Exception as e:
                            logging.warning(f"이미지 삽입 오류 (행 {row}): {e}")

                    # 예약 데이터 서식 지정
                    for col in range(11, ws.max_column - 1):  # K열부터 M열까지
                        cell = ws.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    # 예약률 서식 지정
                    reservation_rate_cell = ws.cell(row=row, column=ws.max_column - 1)  # N열
                    reservation_rate_cell.alignment = Alignment(horizontal='center', vertical='center')

                    # 모든 셀 가운데 정렬
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # 헤더 서식 지정
                for cell in ws[1]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = cell.font.copy(bold=True)

                # 행 높이 조정
                for row in range(2, ws.max_row + 1):
                    ws.row_dimensions[row].height = 80  # 필요에 따라 조정

                # 평균 예약률 계산 및 추가
                try:
                    # "예약률" 열 찾기
                    reservation_rate_col = None
                    for col in range(1, ws.max_column + 1):
                        if ws.cell(row=1, column=col).value == "예약률":
                            reservation_rate_col = col
                            break

                    if reservation_rate_col is not None:
                        total_rate = 0
                        count = 0
                        for row in range(2, ws.max_row + 1):
                            cell_value = ws.cell(row=row, column=reservation_rate_col).value
                            if cell_value:
                                try:
                                    rate = float(cell_value.strip('%')) / 100
                                    total_rate += rate
                                    count += 1
                                except:
                                    pass

                        average_rate = (total_rate / count) if count > 0 else 0

                        # 평균을 위한 새로운 행 추가
                        average_row = ws.max_row + 1
                        label_col = reservation_rate_col - 1  # "예약률" 열 앞 열

                        ws.cell(row=average_row, column=label_col).value = "예약률 전체 평균"
                        ws.cell(row=average_row, column=label_col).alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=average_row, column=label_col).font = ws.cell(row=average_row, column=label_col).font.copy(bold=True)

                        ws.cell(row=average_row, column=reservation_rate_col).value = f"{average_rate:.2%}"
                        ws.cell(row=average_row, column=reservation_rate_col).alignment = Alignment(horizontal='center', vertical='center')

                        logging.info(f"전체 평균 예약률 추가됨: {average_rate:.1%}")
                    else:
                        logging.warning("'예약률' 열을 찾을 수 없습니다.")
                except Exception as e:
                    logging.error(f"평균 예약률 계산 오류: {e}", exc_info=True)

                wb.save(excel_path)
                logging.info(f"엑셀 저장 완료 → {excel_path}")

            else:
                logging.info(f"키워드 '{keyword}'는 데이터 없음 → 엑셀 미생성")

    finally:
        driver.quit()
        logging.info("드라이버가 정상 종료되었습니다.")


# if __name__ == "__main__":
#     # 예시 사용법
#     keyword = "강남"  
#     image_dir = "images"
#     excel_path = "output.xlsx"

#     # 이미지 디렉토리가 존재하지 않으면 생성
#     if not os.path.exists(image_dir):
#         os.makedirs(image_dir)

#     crawl(keyword, image_dir, excel_path, max_sections=5, pages_per_section=10)
